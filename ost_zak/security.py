# -*- coding: utf-8 -*-
"""
Standalone security.py (объединение security.py + chek_3.py) БЕЗ импорта new_project_1_0.py.

Внутри включён минимально необходимый фрагмент логики инициализации из new_project_1_0.py:
- разбор file_config.txt (sheet_names, named_table, unic_params_table, tip_params_table, tip_table_ch и т.д.)
- extract_order_number()

Скрипт сохраняет поведение проверки ИБ:
- excel_read('new', zak) создаёт глобальные списки new_data_serv_*,
- sec_sec() формирует текст отчёта в a_comm,
- при запуске файла напрямую работает интерактивный режим (как раньше).
"""

import os
import numpy as np
from openpyxl import load_workbook
from pathlib import Path

# Логирование отключено для web-режима

def _log(*args, **kwargs):
    """Безопасная замена print() для web-режима."""
    return None

# -----------------------------
# Глобальные переменные/отчёт
# -----------------------------
a_comm = 'Результаты обработки поступивших данных\n'
prom_comm = ''

# -----------------------------
# Минимальная инициализация конфигурации (из new_project_1_0.py)
# -----------------------------
# Эти переменные используются excel_read()
sheet_number = None
sheet_names = None

table_number = []
unic_table_number = []
unic_params_table = []
tip_table_number = []
tip_table_ch = []
tip_params_table = []

named_table = None


def _strip_newline(s):
    if s is None:
        return s
    s = str(s)
    if '\n' in s:
        s = s[:s.find('\n')]
    return s


def _load_config_file_config(base_dir=None):
    """Читает file_config.txt из папки ost_zak (где лежит security.py)."""
    global sheet_number, sheet_names, named_table
    global table_number, unic_table_number, unic_params_table
    global tip_table_number, tip_table_ch, tip_params_table

    # ✅ ВСЕГДА папка, где лежит security.py
    base_dir = Path(__file__).resolve().parent

    path = base_dir / "file_config.txt"
    if not path.exists():
        raise FileNotFoundError(f"Не найден файл конфигурации: {path}")
    # Счётчик строк и позиции ключей
    lines = 0
    sh_line = 0
    sh_n_line = 0

    # Для составления ключей sheet_i
    t_n_s = []
    u_t_n = []
    u_p_t = []
    t_t_n = []
    t_t_ch = []
    t_t_p = []

    # Считываем файл и определяем количество строк
    with open(path, "r", encoding="utf-8") as file_read_params:
        all_lines = file_read_params.readlines()
    lines = len(all_lines)

    # Находим строки-указатели для sheet_number и sheet_names
    for idx, prom_str in enumerate(all_lines):
        if 'sheet_number' in prom_str:
            sh_line = idx + 1
        elif 'sheet_names' in prom_str:
            sh_n_line = idx + 1

    # Считываем значения sheet_number и sheet_names
    for idx, prom_str in enumerate(all_lines):
        if idx == sh_line:
            sheet_number = int(prom_str)
        elif idx == sh_n_line:
            sheet_names = prom_str.split(';\t')
            sheet_names = np.array(sheet_names)
            for j in range(0, len(sheet_names)):
                sheet_names[j] = _strip_newline(sheet_names[j])

    # Составляем ключи параметров для каждого листа
    if sheet_number is not None:
        for j in range(0, int(sheet_number)):
            nn = 'sheet_' + str(j + 1)
            for prom_str in all_lines:
                if nn in prom_str:
                    t_n_s.append('1table_number_' + str(j + 1))
                    u_t_n.append('unic_table_number_' + str(j + 1))
                    u_p_t.append('unic_params_table_' + str(j + 1))
                    t_t_n.append('tip_table_number_' + str(j + 1))
                    t_t_ch.append('tip_table_ch_' + str(j + 1))
                    t_t_p.append('tip_params_table_' + str(j + 1))
                    break

    # Парсим параметры по каждому листу
    if sheet_number is not None:
        for j in range(0, int(sheet_number)):
            table_number_line = 1000000
            unic_table_number_line = 1000000
            unic_params_table_line = 1000000
            tip_table_number_line = 1000000
            tip_table_ch_line = 1000000
            tip_params_table_line = 1000000

            prom_unic_pt = 0

            for idx, prom_str in enumerate(all_lines):
                if t_n_s[j] in prom_str:
                    table_number_line = idx + 1
                elif u_t_n[j] in prom_str:
                    unic_table_number_line = idx + 1
                elif u_p_t[j] in prom_str:
                    unic_params_table_line = idx + 1
                elif t_t_n[j] in prom_str:
                    tip_table_number_line = idx + 1
                elif t_t_ch[j] in prom_str:
                    tip_table_ch_line = idx + 1
                elif t_t_p[j] in prom_str:
                    tip_params_table_line = idx + 1

                # Запись значений (индексы как в оригинале)
                if idx == table_number_line:
                    table_number.append(int(prom_str))
                elif idx == unic_table_number_line:
                    unic_table_number.append(int(prom_str))
                    prom_unic_pt = int(prom_str)
                elif idx >= unic_params_table_line and prom_unic_pt != 0:
                    unic_params_table.append(prom_str.split(';\t'))
                    prom_unic_pt -= 1
                elif idx == tip_table_number_line:
                    tip_table_number.append(int(prom_str))
                elif idx == tip_table_ch_line:
                    tip_table_ch.append(prom_str.split(';\t'))
                elif idx == tip_params_table_line:
                    tip_params_table.append(prom_str.split(';\t'))

    # Определяем named_table
    tt = None
    for idx, prom_str in enumerate(all_lines):
        if 'named_table' in prom_str:
            tt = idx + 1
        if tt is not None and idx == tt:
            named_table = prom_str
            break
    named_table = _strip_newline(named_table)

    # Чистим переносы строк в таблицах параметров
    for i in range(0, len(unic_params_table)):
        for j in range(0, len(unic_params_table[i])):
            unic_params_table[i][j] = _strip_newline(unic_params_table[i][j])

    for i in range(0, len(tip_table_ch)):
        for j in range(0, len(tip_table_ch[i])):
            tip_table_ch[i][j] = _strip_newline(tip_table_ch[i][j])

    # Разворачиваем диапазоны в tip_table_ch (как в оригинале)
    tt_list = []
    dd = None
    for i in range(0, len(tip_table_ch)):
        if tip_table_ch[i] != []:
            for j in range(0, len(tip_table_ch[i])):
                val = tip_table_ch[i][j]
                if '-' in str(val):
                    dd = str(val).split('-')
                    for h in range(int(dd[0]), int(dd[1])):
                        tt_list.append(h)
                elif '' not in str(val) and None not in str(val):
                    tt_list.append(val)
    tip_table_ch[:] = tt_list

    for i in range(0, len(tip_params_table)):
        for j in range(0, len(tip_params_table[i])):
            tip_params_table[i][j] = _strip_newline(tip_params_table[i][j])


def extract_order_number(workbook_path: str) -> str:
    """Извлекает номер заказа с листа 'Титул' (как в new_project_1_0.py)."""
    workbook = load_workbook(filename=workbook_path)
    tip = 0
    u_tip = 3
    order_number = ''
    if "Титул" in workbook.sheetnames:
        sheet_title = workbook["Титул"]
        target_phrase = "ЗАКАЗ НА ОКАЗАНИЕ УСЛУГ №"
        for row in sheet_title.iter_rows(values_only=True):
            for cell in row:
                if target_phrase in str(cell):
                    tip = tip + 1
                if tip > 0 and u_tip > 0:
                    tip = tip + 1
                    if cell is not None and str(cell) != target_phrase:
                        order_number = order_number + str(cell).strip() + ' '
                        u_tip = u_tip - 1
            order_number = order_number.replace(target_phrase, '').strip()
            if order_number:
                break
    return order_number


# Загружаем конфигурацию при импорте модуля (как это делалось ранее через new_project_1_0.py)
_load_config_file_config()


# -----------------------------
# Excel reader (из security.py/chek_3.py)
# -----------------------------
def excel_read(new_old: str, zak: str) -> int:
    """Читает Excel-файл заказа и раскладывает строки таблиц в глобальные переменные new_data_serv_*.

    Возвращает qs (0 если ок, 7 если нет нужных листов).
    """
    zk_type = new_old
    qs = 0

    wb = load_workbook(filename=zak, data_only=True)
    sheets = wb.sheetnames

    global a_comm
    a_comm += extract_order_number(zak) + '\n'

    ws = []
    for i in sheets:
        for j in range(0, len(sheet_names)):
            if str(sheet_names[j]) in i:
                ws.append(wb[i])

    if ws == []:
        _log('Данный файл не содержит листов указанных в конфигурационном файле')
        qs = 7

    _log('Начата обработка поступившего файла')

    for i in range(0, len(ws)):
        rows = ws[i].max_row
        columns = ws[i].max_column
        tables_rows = []

        for ii in range(1, rows):
            for jj in range(1, columns):
                if str(named_table) in str(ws[i].cell(row=ii, column=jj).value):
                    tables_rows.append(ii)

        for k in range(0, len(tables_rows)):
            col_params = []
            row_params = None
            chek_l = 10000000000000

            end_row = tables_rows[k + 1] if k != len(tables_rows) - 1 else rows

            for ii in range(tables_rows[k], end_row):
                new_line = []
                for jj in range(1, columns):
                    if named_table in str(ws[i].cell(row=ii, column=jj).value):
                        name = str(ws[i].cell(row=ii, column=jj).value)
                        nn = name.find(named_table)
                        if name[nn + len(named_table) + 2] != '.':
                            name = name[nn + len(named_table) + 1:nn + len(named_table) + 2 + 1]
                        else:
                            name = name[nn + len(named_table) + 1]

                    for m in range(0, len(unic_params_table)):
                        if int(name) == int(unic_params_table[m][0]):
                            chek_l = m
                    if chek_l >= len(unic_params_table):
                        chek_l = -1

                    if chek_l != -1 and chek_l != len(unic_params_table):
                        for m in range(1, len(unic_params_table[chek_l])):
                            if str(unic_params_table[chek_l][m]) in str(ws[i].cell(row=ii, column=jj).value):
                                if jj not in col_params:
                                    col_params.append(jj)
                                    row_params = ii
                        for m in range(0, len(col_params)):
                            if jj == int(col_params[m]) and ii > row_params:
                                new_line.append(ws[i].cell(row=ii, column=jj).value)
                    else:
                        for m in tip_table_ch:
                            if int(name) == int(m):
                                chek_l = m
                        if chek_l != 10000000000000:
                            for m in range(0, len(tip_params_table[i])):
                                if str(tip_params_table[i][m]) in str(ws[i].cell(row=ii, column=jj).value):
                                    if jj not in col_params:
                                        col_params.append(jj)
                                        row_params = ii
                            for m in range(0, len(col_params)):
                                if jj == int(col_params[m]) and ii > row_params:
                                    new_line.append(ws[i].cell(row=ii, column=jj).value)

                q = str(i + 1) + '_' + str(name)
                qq = '{0}_data_serv_{1}'.format(zk_type, q)

                if str(qq) not in globals():
                    if new_line != [] and new_line[1] is not None and new_line[0] is not None:
                        exec('%s=[]' % qq, globals())
                        exec('%s.append(new_line)' % qq)
                else:
                    if new_line != [] and new_line[1] is not None and new_line[0] is not None:
                        exec('%s.append(new_line)' % qq)

    _log('Обработка файла завершена')
    try:
        wb.close()
    except Exception:
        pass
    return qs


# -----------------------------
# Проверки ИБ (из chek_3.py)
# -----------------------------
def count_211(data_set: list):
    res = 0
    for i in range(0, len(data_set)):
        if 'услуга 2.1.1' in str(data_set[i]) and 'PROD' in str(data_set[i]) and (
                'Заказанная услуга' in str(data_set[i]) or 'Новая услуга' in str(data_set[i])):
            res = res + 1
    return res


def count_21827(data_set: list, numn_name, numn_status, numn_count):
    res = 0
    nn = 0
    rows_list = []
    for i in range(0, len(data_set)):
        if '2.18.27' in str(data_set[i][numn_name]) and (
                'Заказанная' in str(data_set[i][numn_status]) or 'Новая' in str(data_set[i][numn_status])):
            nn = nn + 1
            rows_list.append(i)
    if nn == 1:
        res = int(data_set[rows_list[0]][numn_count])
    elif nn > 0 and nn != 1:
        for i in range(0, len(rows_list)):
            res = res + int(data_set[rows_list[i]][numn_count])
    return res


def chek_218(data_set: list, numn_name, numn_status):
    res = []
    for i in range(0, len(data_set)):
        if '2.18.1)' in str(data_set[i][numn_name]) or '2.18.2)' in str(data_set[i][numn_name]) or '2.18.3)' in str(
                data_set[i][numn_name]) or '2.18.4' in str(data_set[i][numn_name]) or '2.18.5' in str(
                data_set[i][numn_name]) or '2.18.6' in str(data_set[i][numn_name]) or '2.18.7' in str(
                data_set[i][numn_name]) or '2.18.8' in str(data_set[i][numn_name]) or '2.18.9' in str(
                data_set[i][numn_name]) or '2.18.10' in str(data_set[i][numn_name]) or '2.18.11' in str(
                data_set[i][numn_name]) or '2.18.16' in str(data_set[i][numn_name]) or '2.18.17' in str(
                data_set[i][numn_name]) or '2.18.18' in str(data_set[i][numn_name]) or '2.18.19' in str(
                data_set[i][numn_name]) or '2.18.20' in str(data_set[i][numn_name]) or '2.18.21' in str(
                data_set[i][numn_name]) or '2.18.22' in str(data_set[i][numn_name]) or '2.18.23' in str(
                data_set[i][numn_name]) or '2.18.24' in str(data_set[i][numn_name]) or '2.18.25' in str(
                data_set[i][numn_name]) or '2.18.26' in str(data_set[i][numn_name]) or '2.18.28' in str(
                data_set[i][numn_name]) or '2.18.29' in str(data_set[i][numn_name]) or '2.18.30' in str(
                data_set[i][numn_name]) or '2.18.31' in str(data_set[i][numn_name]) or '2.18.32' in str(
                data_set[i][numn_name]) or '2.18.33' in str(data_set[i][numn_name]) or '2.18.36' in str(
                data_set[i][numn_name]) or '2.18.37' in str(data_set[i][numn_name]):
            if 'Оказанная' not in str(data_set[i][numn_status]) and 'Аннуляция' not in str(data_set[i][numn_status]):
                res.append(i)
    return res


def count_channels(data_set: list, numn_name, numn_type, numn_status, numn_count):
    varr = []
    varr2 = []
    numn = None
    for i in range(0, len(data_set)):
        if 'Канал связи L' in str(data_set[i][numn_name]) and (
                'Заказанная услуга' in str(data_set[i][numn_status]) or 'Новая услуга' in str(
                data_set[i][numn_status])):
            var_get = [data_set[i][numn_name], data_set[i][numn_type], data_set[i][numn_count]]
            varr.append(var_get)
    for i in range(0, len(varr)):
        var_get = varr[i][0].split(' ')
        varr2.append(var_get)
    for i in range(0, len(varr2)):
        for j in range(0, len(varr2[i])):
            if 'L' in str(varr2[i][j]) and numn == None:
                numn = j
    varr3 = []
    for i in range(0, len(varr)):
        var_get = [varr2[i][numn + 1], varr[i][1], varr[i][2]]
        varr3.append(var_get)

    my_comm = None
    varr = []
    for i in range(0, (len(varr3) - 1)):
        swit = None
        if 'Основной' in str(varr3[i][1]) and 'Резервный' in str(varr3[i + 1][1]):
            if varr3[i][0] == varr3[i + 1][0] and varr3[i][2] == varr3[i + 1][2]:
                swit = True
                var_get = [varr3[i][0], varr3[i][2]]
                varr.append(var_get)
        if swit == True:
            break
        else:
            if my_comm == None:
                my_comm = 'Количество и тип основных и резервных каналов не соответствуют.'
            else:
                break
    varr2 = []
    comp_table = []
    for i in range(0, len(varr)):
        comp_table.append(False)
    compp = np.array(comp_table)

    if (len(varr) - 1) != 0:
        for i in range(0, (len(varr) - 1)):
            for j in range(i + 1, len(varr)):
                if varr[i][0] == varr[j][0] and compp[j] == False:
                    scount = int(varr[i][1]) + int(varr[j][1])
                    var_get = [varr[i][0], scount]
                    varr2.append(var_get)
                    compp[j] = True
    else:
        i = 0
        var_get = [varr[i][0], varr[i][1]]
        varr2.append(var_get)

    varr = []
    varm = np.array(varr2)
    comp_table = []

    if (len(varm) - 1) != 0:
        for i in range(0, len(varm)):
            comp_table.append(False)
        compp = np.array(comp_table)
        for i in range(0, len(varm)):
            if int(varm[i, 0]) == 100:
                varm[i, 0] = 1
        for i in range(0, len(varm) - 1):
            for j in range(0, len(varm)):
                # Сохранено как в исходнике: compp == False (а не compp[j] == False)
                if varm[i, 0] == varm[j, 0] and compp == False:
                    scount = int(varm[i, 1]) + int(varm[j, 1])
                    var_get = [varm[i, 0], scount]
                    varr.append(var_get)
                    compp[j] = True
    else:
        i = 0
        if int(varm[i, 0]) == 100:
            varm[i, 0] = 1
        var_get = [varm[i, 0], varm[i, 1]]
        varr.append(var_get)

    res = varr
    return res


def count_skzi(data_set: list, numn_name, numn_status, numn_count):
    varr = []
    varr2 = []
    numn = None
    for i in range(0, len(data_set)):
        if 'Комплект средств криптографической защиты информации уровня L2/L3' in str(data_set[i][numn_name]) and (
                'Заказанная' in str(data_set[i][numn_status]) or 'Новая' in str(data_set[i][numn_status])):
            var_get = [data_set[i][numn_name], data_set[i][numn_count]]
            varr.append(var_get)
    for i in range(0, len(varr)):
        var_get = varr[i][0].split(' ')
        varr2.append(var_get)
    for i in range(0, len(varr2)):
        for j in range(0, len(varr2[i])):
            if 'способностью' in str(varr2[i][j]) and numn == None:
                numn = j
    varr3 = []
    for i in range(0, len(varr)):
        var_get = [varr2[i][numn + 1], varr[i][1]]
        varr3.append(var_get)

    varr2 = []
    comp_table = []
    for i in range(0, len(varr)):
        comp_table.append(False)
    compp = np.array(comp_table)

    if (len(varr3) - 1) != 0:
        for i in range(0, (len(varr3) - 1)):
            for j in range(i + 1, len(varr3)):
                if varr3[i][0] == varr3[j][0] and compp[j] == False:
                    scount = int(varr3[i][1]) + int(varr3[i][1])
                    var_get = [varr3[i][0], scount]
                    varr2.append(var_get)
                    compp[j] = True
    else:
        i = 0
        var_get = [varr3[i][0], varr3[i][1]]
        varr2.append(var_get)

    res = varr2
    return res


def count_podd(data_set_1: list, numn_name_1, numn_stend_1, numn_status_1, numn_type_os_1, data_set_2: list,
               numn_name_2, numn_stend_2, numn_status_2):
    varr = []

    for i in range(0, len(data_set_1)):
        if 'услуга 1.18' in str(data_set_1[i][numn_name_1]) and int(data_set_1[i][numn_type_os_1]) == 1 and (
                'Новая' in str(data_set_1[i][numn_status_1]) or 'Заказанная' in str(data_set_1[i][numn_status_1])):
            var_get = [data_set_1[i][numn_stend_1], 1]
            varr.append(var_get)

    for i in range(0, len(data_set_2)):
        if 'услуга 2.19.2' in str(data_set_2[i][numn_name_2]) and (
                'Новая' in str(data_set_2[i][numn_status_2]) or 'Заказанная' in str(data_set_2[i][numn_status_2])):
            var_get = [data_set_2[i][numn_stend_2], 1]
            varr.append(var_get)

    varr2 = np.array(varr)
    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if varr2[i, 0] == varr2[j, 0]:
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[j, 1])
                    varr2[j, 1] = 0
    varr = []
    for i in range(0, len(varr2)):
        if int(varr2[i, 1]) != 0:
            var_get = [varr2[i, 0], varr2[i, 1]]
            varr.append(var_get)
    res = varr
    return res


def count_sec_podd(data_set: list, numn_name, numn_stend, numn_scount):
    varr = []
    for i in range(0, len(data_set)):
        if 'услуга 2.18.35' in str(data_set[i][numn_name]) and (
                'Заказанная услуга' in str(data_set[i]) or 'Новая услуга' in str(data_set[i])):
            var_get = [data_set[i][numn_stend], data_set[i][numn_scount]]
            varr.append(var_get)

    varr2 = np.array(varr)

    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if varr2[i, 0] == varr2[j, 0]:
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[j, 1])
                    varr2[j, 1] = 0
    varr = []
    for i in range(0, len(varr2)):
        if int(varr2[i, 1]) != 0:
            var_get = [varr2[i, 0], varr2[i, 1]]
            varr.append(var_get)

    res = varr
    return res


def count_follow(data_set: list, numn_name, numn_status, numn_scount):
    varr = []
    for i in range(0, len(data_set)):
        if ('3.1.5' in str(data_set[i][numn_name]) or '3.1.6' in str(data_set[i][numn_name]) or '3.1.18' in str(
                data_set[i][numn_name]) or '3.1.25' in str(data_set[i][numn_name])) and (
                'Новая' in str(data_set[i][numn_status]) or 'Заказанная' in str(data_set[i][numn_status])):
            var_get = [data_set[i][numn_name], data_set[i][numn_scount]]
            varr.append(var_get)
    varr2 = np.array(varr)
    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if varr2[i, 0] == varr2[j, 0]:
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[j, 1])
                    varr2[j, 1] = 0
    varr = []
    for i in range(0, len(varr2)):
        if int(varr2[i, 1]) != 0:
            var_get = [varr2[i, 0], varr2[i, 1]]
            varr.append(var_get)
    res = varr
    return res


def name_skzi(data_set: list, numn_name, numn_status, numn_count):
    varr = []
    for i in range(0, len(data_set)):
        if 'Комплект средств криптографической защиты информации уровня L2/L3' in str(data_set[i][numn_name]) and (
                'Заказанная' in str(data_set[i][numn_status]) or 'Новая' in str(data_set[i][numn_status])):
            var_get = [data_set[i][numn_name], data_set[i][numn_count]]
            varr.append(var_get)

    varr2 = np.array(varr)
    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if varr2[i, 0] == varr2[j, 0]:
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[i, 1])
                    varr2[j, 1] = 0
    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if ('2.18.12' in str(varr2[i, 0]) and '2.18.13' in str(varr2[j, 0])) or (
                        '2.18.15' in str(varr2[i, 0]) and '2.18.16' in str(varr2[j, 0])):
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[j, 1])
                    varr2[j, 0] = 0
                    varr2[j, 1] = 0

    varr = []
    for i in range(0, len(varr2)):
        if int(varr2[i, 1]) != 0:
            var_get = [varr2[i, 0], varr2[i, 1]]
            varr.append(var_get)
    res = varr
    return res


def count_sig(data_set: list, numn_name, numn_status, numn_count):
    varr = []
    for i in range(0, len(data_set)):
        if '2.18.34' in str(data_set[i][numn_name]) and (
                'Заказанная' in str(data_set[i][numn_status]) or 'Новая' in str(data_set[i][numn_status])):
            var_get = [data_set[i][numn_name], data_set[i][numn_count]]
            varr.append(var_get)

    varr2 = np.array(varr)
    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if varr2[i, 0] == varr2[j, 0]:
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[i, 1])
                    varr2[j, 1] = 0

    varr = []
    for i in range(0, len(varr2)):
        if int(varr2[i, 1]) != 0:
            var_get = [varr2[i, 0], varr2[i, 1]]
            varr.append(var_get)

    res = varr
    return res


def count_ip(data_set: list, numn_name, numn_status, numn_count):
    varr = []
    for i in range(0, len(data_set)):
        if '2.16.1' in str(data_set[i][numn_name]) and ('Заказанная' in str(data_set[i][numn_status]) or 'Новая' in str(
                data_set[i][numn_status])) and 'PROD' in str(data_set[i]):
            var_get = [data_set[i][numn_name], data_set[i][numn_count]]
            varr.append(var_get)

    varr2 = np.array(varr)
    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if varr2[i, 0] == varr2[j, 0]:
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[j, 1])
                    varr2[j, 1] = 0

    varr = []
    for i in range(0, len(varr2)):
        if int(varr2[i, 1]) != 0:
            var_get = [varr2[i, 0], varr2[i, 1]]
            varr.append(var_get)

    res = varr
    return res


def count_32(data_set: list, numn_name, numn_status, numn_count):
    varr = []
    for i in range(0, len(data_set)):
        if '3.2' in str(data_set[i][numn_name]) and (
                'Заказанная' in str(data_set[i][numn_status]) or 'Новая' in str(data_set[i][numn_status])):
            var_get = [data_set[i][numn_name], data_set[i][numn_count]]
            varr.append(var_get)

    varr2 = np.array(varr)
    if len(varr2) != 1:
        for i in range(0, len(varr2) - 1):
            for j in range(i + 1, len(varr2)):
                if varr2[i, 0] == varr2[j, 0]:
                    varr2[i, 1] = int(varr2[i, 1]) + int(varr2[j, 1])
                    varr2[j, 1] = 0

    varr = []
    for i in range(0, len(varr2)):
        if int(varr2[i, 1]) != 0:
            var_get = [varr2[i, 0], varr2[i, 1]]
            varr.append(var_get)

    res = varr
    return res


def chek_31(data_set: list, numn_name, numn_status):
    res = []
    for i in range(0, len(data_set)):
        if '3.1.5' not in str(data_set[i][numn_name]) and '3.1.6' not in str(
                data_set[i][numn_name]) and '3.1.18' not in str(data_set[i][numn_name]) and '3.1.25' not in str(
                data_set[i][numn_name]):
            if 'Оказанная' not in str(data_set[i][numn_status]) and 'Аннуляция' not in str(data_set[i][numn_status]):
                res.append(i)
    return res


def sec_sec():
    con = 1

    # Безопасные значения по умолчанию (чтобы не падать на len(None))
    vms = 0
    an_vir = 0
    skzi = []
    sec_podd = []
    skzi_name = []
    sig = []
    n_sec_serv = []
    chanels = []
    podd = []
    follow_sec = []
    serv_32 = []
    ip_addr = []

    global a_comm
    start_len = len(a_comm)
    _log('Начата проверка поступившего файла на услуги ИБ')

    if 'new_data_serv_1_1' in globals():
        vms = count_211(new_data_serv_1_1)

    a_comm += "Таблица 18\n"
    if 'new_data_serv_2_18' in globals():
        an_vir = count_21827(new_data_serv_2_18, 4, 5, 8)
        skzi = count_skzi(new_data_serv_2_18, 4, 5, 8)
        sec_podd = count_sec_podd(new_data_serv_2_18, 4, 3, 8)
        skzi_name = name_skzi(new_data_serv_2_18, 4, 3, 8)
        sig = count_sig(new_data_serv_2_18, 4, 3, 8)
        n_sec_serv = chek_218(new_data_serv_2_18, 4, 3)

    if (vms is not None) and (an_vir is not None):
        if int(vms) != int(an_vir):
            a_comm = a_comm + 'В заказе присутсвует противоречие количества заказываемых ВМ на PROD стенд и числа средств антивирусной защиты. Заказывается ' + str(
                vms) + ' ВМ и при этом ' + str(
                an_vir) + ' средств антивирусной защиты. Прошу сравнить количество. Отмечаю, что для каждой ВМ на PROD требуется заказ средства антивирусной защиты.' + ''

    if n_sec_serv != []:
        for i in range(0, len(n_sec_serv)):
            prom_comm = 'В строке ' + str(new_data_serv_2_18[n_sec_serv[i]][0]) + ' присутствует ' + str(
                new_data_serv_2_18[n_sec_serv[i]][2]) + '. Данная услуга не заказывается Потребителем услуг. Необходимо исключить'
            a_comm = a_comm + prom_comm + ''

    if 'new_data_serv_2_15' in globals():
        chanels = count_channels(new_data_serv_2_15, 2, 3, 4, 6)
    if (skzi is not None) and (chanels is not None):
        if len(skzi) == len(chanels):
            comppp = []
            for zz in range(0, len(skzi)):
                comppp.append(False)
            comp_table = np.array(comppp)
            for i in range(0, len(skzi)):
                prom_comm = None
                for j in range(0, len(skzi)):
                    if skzi[i][0] == chanels[j][0] and comp_table[i] == False:
                        if int(skzi[i][1]) == int(chanels[j][1]):
                            comp_table[i] = True
                            break
                        else:
                            prom_comm = 'Не соответствует количество заказываемых каналов связи и средств криптографической защиты. Каналов связи - ' + str(
                                chanels[j][1]) + ', а средств криптографической защиты информации пропускной способностью ' + str(
                                skzi[i][0]) + ' Гбит/с - ' + str(skzi[i][1])
                            comp_table[i] = True
                if comp_table[i] == False:
                    prom_comm = 'Отсутсвуют каналы связи, для которых заказываются средства криптографической защиты информации пропускной способностью ' + str(
                        skzi[i][0]) + ' Гбит/с в количестве ' + str(skzi[i][1]) + ' штук.'

                if prom_comm != None:
                    a_comm = a_comm + prom_comm + ''
        elif len(skzi) > len(chanels) and len(chanels) != 0:
            skzi_matr = np.array(skzi)
            a_comm = a_comm + 'Количество заказываемых типов средств криптографической защиты информации L2/L3 больше количества заказанных каналов - прошу перепроверить необходимость средств криптографической защиты информации пропускной способностью ' + str(
                skzi_matr[:, 0]) + ' Гбит/с заказываются в количестве ' + str(skzi_matr[:, 1]) + ' соответственно.'
            a_comm = a_comm + ''
        elif len(skzi) < len(chanels) and len(skzi) != 0:
            skzi_matr = np.array(skzi)
            a_comm = a_comm + 'Количество заказываемых типов средств криптографической защиты информации L2/L3 меньше количества заказанных каналов - прошу перепроверить необходимость средств криптографической защиты информации пропускной способностью ' + str(
                skzi_matr[:, 0]) + ' Гбит/с заказываются в количестве ' + str(skzi_matr[:, 1]) + ' соответственно.'
            a_comm = a_comm + ''
        elif len(skzi) == 0 and len(chanels) != 0:
            a_comm = a_comm + 'Не заказываются средcтва криптографической защиты информации, при этом заказываются каналы связи - прошу проверить.'
            a_comm = a_comm + ''
        elif len(skzi) != 0 and len(chanels) == 0:
            a_comm = a_comm + 'Заказываются средcтва криптографической защиты информации, при этом не заказываются каналы связи - прошу проверить.'
            a_comm = a_comm + ''

    if 'new_data_serv_1_1' in globals() and 'new_data_serv_2_19' in globals():
        podd = count_podd(new_data_serv_1_1, 3, 7, 8, 18, new_data_serv_2_19, 2, 3, 4)

    comppp = []
    comppp_1 = []

    for i in range(0, len(podd)):
        comppp.append(False)
    for i in range(0, len(sec_podd)):
        comppp_1.append(False)
    comp_table = np.array(comppp)
    comp_table_1 = np.array(comppp_1)

    if len(podd) != 0 and len(sec_podd) != 0:
        for i in range(0, len(podd)):
            for j in range(0, len(sec_podd)):
                if podd[i][0] == sec_podd[j][0]:
                    if podd[i][1] == sec_podd[j][1]:
                        comp_table[i] = True
                        comp_table_1[j] = True
                        break
                    else:
                        prom_comm = 'Заказанное количество агентов ПОДД и средств криптографической защиты агентов ПОДД не соответствует друг другу. Агентов ПОДД заказывается - ' + str(
                            podd[i][1]) + ', а средств защиты к ним - ' + str(sec_podd[j][1]) + ' для стенда ' + str(
                            podd[i][0])
                        comp_table[i] = True
                        comp_table_1[j] = True
                        a_comm = a_comm + prom_comm + ''
        for i in range(0, len(podd)):
            if comp_table[i] == False:
                prom_comm = 'Для заказываемых на стенде ' + str(podd[i][0]) + ' Агентов ПОДД не заказываются средства криптографической защиты информации. Необходимо скорректировать. В заказе присутсвует ' + str(
                    podd[i][1]) + ' Агентов ПОДД для данного стенда.'
                a_comm = a_comm + prom_comm + ''
        for i in range(0, len(sec_podd)):
            if comp_table_1[i] == False:
                prom_comm = 'Для заказываемых на стенде ' + str(sec_podd[i][0]) + ' средств криптографической защиты информации не заказываются Агенты ПОДД. Необходимо скорректировать. В заказе присутсвует ' + str(
                    sec_podd[i][1]) + ' средств криптографической защиты информации Агентов ПОДД для данного стенда.'
                a_comm = a_comm + prom_comm + ''
    elif len(podd) == 0 and len(sec_podd) != 0:
        sec_podd_arr = np.array(sec_podd)
        prom_comm = 'В заказе отсутствуют Агенты ПОДД, при этом заказываются средства криптографической защиты Агентов ПОДД для ' + str(
            sec_podd_arr[:, 0]) + ' в количестве ' + str(sec_podd_arr[:, 1]) + ' соответственно.'
        a_comm = a_comm + prom_comm + ''
    elif len(sec_podd) == 0 and len(podd) != 0:
        podd_arr = np.array(podd)
        prom_comm = 'В заказе отсутствуют средства криптографической защиты Агентов ПОДД, при этом заказываются Агенты ПОДД для ' + str(
            podd_arr[:, 0]) + ' в количестве ' + str(podd_arr[:, 1]) + ' соответственно.'
        a_comm = a_comm + prom_comm + ''

    if 'new_data_serv_3_20' in globals():
        follow_sec = count_follow(new_data_serv_3_20, 2, 3, 5)
        n_sec_serv = chek_31(new_data_serv_3_20, 2, 3)

    a_comm += "Таблица 20\n"

    if n_sec_serv != []:
        for i in range(0, len(n_sec_serv)):
            prom_comm = 'В строке ' + str(new_data_serv_3_20[n_sec_serv[i]][0]) + ' присутствует ' + str(
                new_data_serv_3_20[n_sec_serv[i]][2]) + '. Данная услуга не заказывается Потребителем услуг. Необходимо исключить'
            a_comm = a_comm + prom_comm + ''

    if len(follow_sec) != 0:
        comppp = []
        comp_an_vir = None
        comp_sig = None

        for i in range(0, len(follow_sec)):
            comppp.append(False)
        comp_table = np.array(comppp)

        if len(skzi_name) != 0:
            comppp_1 = []
            for ii in range(0, len(skzi_name)):
                comppp_1.append(False)
            comp_table_1 = np.array(comppp_1)

        for i in range(0, len(follow_sec)):

            if len(skzi_name) != 0:
                for j in range(0, len(skzi_name)):
                    if '3.1.5' in str(follow_sec[i][0]) and 'тип 1' in str(skzi_name[j][0]) and comp_table[
                        i] == False and comp_table_1[j] == False:
                        if int(follow_sec[i][1]) != int(skzi_name[j][1]):
                            prom_comm = 'Количество заказываемых комплектов средств криптографической защиты информации уровня L2/L3 Тип 1 не совпадает с количеством услуг по их сопровождению. Сопровождение заказывается - ' + str(
                                follow_sec[i][1]) + ' , при этом комплетков средств заказывается - ' + str(
                                skzi_name[j][1])
                            a_comm = a_comm + prom_comm + ''
                            comp_table[i] = True
                            comp_table_1[j] = True
                        else:
                            comp_table[i] = True
                            comp_table_1[j] = True
                    elif '3.1.6' in str(follow_sec[i][0]) and 'тип 2' in str(skzi_name[j][0]) and comp_table[
                        i] == False and comp_table_1[j] == False:
                        if int(follow_sec[i][1]) != int(skzi_name[j][1]):
                            prom_comm = 'Количество заказываемых комплектов средств криптографической защиты информации уровня L2/L3 Тип 2 не совпадает с количеством услуг по их сопровождению. Сопровождение заказывается - ' + str(
                                follow_sec[i][1]) + ' , при этом комплетков средств заказывается - ' + str(
                                skzi_name[j][1])
                            a_comm = a_comm + prom_comm + ''
                            comp_table[i] = True
                            comp_table_1[j] = True
                        else:
                            comp_table[i] = True
                            comp_table_1[j] = True

            else:
                if ('3.1.5' in str(follow_sec[i][0]) or '3.1.6' in str(follow_sec[i][0])) and comp_table[i] == False:
                    prom_comm = 'Заказываются услуги по сопровождению комплектов средств криптографической защиты информации уровня L2/L3 при этом сами комплекты не заказываются. Прошу перепроверить, заказывается ' + str(
                        follow_sec[i][0])
                    a_comm = a_comm + prom_comm + ''

            if an_vir != 0:
                if '3.1.18' in str(follow_sec[i][0]) and comp_table[i] == False:
                    if int(follow_sec[i][1]) != int(an_vir) and an_vir != 0:
                        prom_comm = 'Количество заказываемых средств антивирусной защиты не совпадает с количеством услуг по их сопровождению. Количество средств антивирусной защиты - ' + str(
                            an_vir) + ', а количество услуг по их сопровождению - ' + str(follow_sec[i][1])
                        a_comm = a_comm + prom_comm + ''
                        comp_table[i] = True
                        comp_an_vir = True
                    else:
                        comp_table[i] = True
                        comp_an_vir = True
            if sig != []:
                if '3.1.25' in str(follow_sec[i][0]) and comp_table[i] == False:
                    if int(follow_sec[i][1]) != int(sig[0][1]):
                        prom_comm = 'Количество заказываемых средств автоматизированного формирования и проверки электронной подписи КС2 не совпадает с количеством услуг по их сопровождению. Количество средств подписания - ' + str(
                            sig[0][1]) + ', а количество услуг по их сопровождению - ' + str(follow_sec[i][1])
                        a_comm = a_comm + prom_comm + ''
                        comp_table[i] = True
                        comp_sig = True
                    else:
                        comp_table[i] = True
                        comp_sig = True
        for i in range(0, len(follow_sec)):
            if comp_table[i] == False:
                prom_comm = 'Для ' + str(follow_sec[i][0]) + ' не найдены комплекты средств, подлежащие сопровождению.'
                a_comm = a_comm + prom_comm + ''
                comp_table[i] = True
        for i in range(0, len(skzi_name)):
            if comp_table_1[i] == False:
                if 'тип 1' in str(skzi_name[i][0]):
                    prom_comm = 'Для комплектов средств криптографической защиты ' + 'тип 1 в количестве - ' + str(
                        skzi_name[i][1]) + ' штук не найдены комплекты средств, подлежащие сопровождению.'
                    a_comm = a_comm + prom_comm + ''
                    comp_table_1[i] = True
                elif 'тип 2' in str(skzi_name[i][0]):
                    prom_comm = 'Для комплектов средств криптографической защиты ' + 'тип 2 в количестве - ' + str(
                        skzi_name[i][1]) + ' штук не найдены комплекты средств, подлежащие сопровождению.'
                    a_comm = a_comm + prom_comm + ''
                    comp_table_1[i] = True
        if comp_an_vir == None and int(an_vir) != 0:
            prom_comm = 'Для средств антивирусной защиты в количестве - ' + str(
                an_vir) + ' штуки не найдены услуги по их сопровождению.'
            a_comm = a_comm + prom_comm + ''
            comp_an_vir = True
        if comp_sig == None and sig != []:
            prom_comm = 'Для ' + str(sig[0][0]) + ' в количестве - ' + str(
                sig[0][1]) + ' штуки не найдены услуги по их сопровождению.'
            a_comm = a_comm + prom_comm + ''
            comp_sig = True
    else:
        comp_an_vir = None
        comp_sig = None
        if (skzi is not None):
            if len(skzi_name) != 0:
                comppp_1 = []
                for ii in range(0, len(skzi_name)):
                    comppp_1.append(False)
                comp_table_1 = np.array(comppp_1)

            for i in range(0, len(skzi_name)):
                if comp_table_1[i] == False:
                    if 'тип 1' in str(skzi_name[i][0]):
                        prom_comm = 'Для комплектов средств криптографической защиты ' + 'тип 1 в количестве - ' + str(
                            skzi_name[i][1]) + ' штук не найдены комплекты средств, подлежащие сопровождению.'
                        a_comm = a_comm + prom_comm + ''
                        comp_table_1[i] = True
                    elif 'тип 2' in str(skzi_name[i][0]):
                        prom_comm = 'Для комплектов средств криптографической защиты ' + 'тип 2 в количестве - ' + str(
                            skzi_name[i][1]) + ' штук не найдены комплекты средств, подлежащие сопровождению.'
                        a_comm = a_comm + prom_comm + ''
                        comp_table_1[i] = True
            if comp_an_vir == None and int(an_vir) != 0:
                prom_comm = 'Для средств антивирусной защиты в количестве - ' + str(
                    an_vir) + ' штук не найдены услуги по их сопровождению.'
                a_comm = a_comm + prom_comm + ''
                comp_an_vir = True
            if comp_sig == None and sig != []:
                prom_comm = 'Для ' + str(sig[0][0]) + ' в количестве - ' + str(
                    sig[0][1]) + ' штук не найдены услуги по их сопровождению.'
                a_comm = a_comm + prom_comm + ''
                comp_sig = True

    if 'new_data_serv_2_16' in globals():
        if 'DEV' in str(new_data_serv_2_16) or 'TEST' in str(new_data_serv_2_16):
            prom_comm = 'Обращаю внимание, что публичные IP-адреса не могут быть заказаны для DEV/TEST стендов.'
            a_comm = a_comm + prom_comm + ''
            ip_addr = count_ip(new_data_serv_2_16, 2, 5, 7)

    if 'new_data_serv_3_21' in globals():
        serv_32 = count_32(new_data_serv_3_21, 2, 3, 5)
        a_comm = a_comm + 'Таблица 21'

    if serv_32 != []:
        comppp = []
        for i in range(0, len(serv_32)):
            comppp.append(False)
        comp_table = np.array(comppp)

        for i in range(0, len(serv_32)):
            if '3.2.1' in str(serv_32[i][0]) or '3.2.2' in str(serv_32[i][0]) or '3.2.3' in str(serv_32[i][0]):
                if ip_addr != []:
                    comppp_1 = []
                    for j in range(0, len(ip_addr)):
                        comppp_1.append(False)
                    comp_table_1 = np.array(comppp_1)
                    if '3.2.3' in str(serv_32[i][0]) and int(serv_32[i][1]) != int(ip_addr[0][1]) and comp_table[
                        i] == False:
                        prom_comm = 'Количество заказываемых ' + str(
                            serv_32[i][0]) + ' не совпадает с количеством заказываемых ' + str(
                            ip_addr[0][0]) + '. Количество ' + str(serv_32[i][0]) + ' - ' + str(
                            serv_32[i][1]) + ', а количество ' + str(ip_addr[0][0]) + ' - ' + str(ip_addr[0][1])
                        a_comm = a_comm + prom_comm + ''
                        comp_table[i] = True
                        comp_table_1[j] = True
                    elif '3.2.3' in str(serv_32[i][0]) and int(serv_32[i][1]) == int(ip_addr[0][1]) and comp_table[
                        i] == False:
                        comp_table[i] = True
                        comp_table_1[j] = True
                else:
                    prom_comm = 'Невозможно заказать ' + str(
                        serv_32[i][0]) + ' ввиду отсутствия заказываемых публичных ip-адресов.'
                    a_comm = a_comm + prom_comm + ''
                    comp_table[i] = True
            elif '3.2.5' in str(serv_32[i][0]):
                if vms != None and vms != 0:
                    comp_vms = None
                    sec_vms_1 = vms // 50
                    if vms % 50 != 0:
                        sec_vms_1 += 1
                    if int(serv_32[i][1]) != int(sec_vms_1) and comp_table[i] == False:
                        prom_comm = 'Неверное количество ' + str(
                            serv_32[i][0]) + '. Для заказываемого количества ВМ - ' + str(
                            vms) + ' необходимо заказывать - ' + str(
                            sec_vms_1) + ' штук данной услуги. Каждая единица услуги имеет квант 50 ВМ.'
                        a_comm = a_comm + prom_comm + ''
                        comp_table[i] = True
                        comp_vms = True
                    elif int(serv_32[i][1]) == int(sec_vms_1) and comp_table[i] == False:
                        comp_table[i] = True
                        comp_vms = True
                else:
                    prom_comm = str(serv_32[i][0]) + ' заказывается только при наличии ВМ в PROD. В заказе отсуствуют ВМ в PROD. Необходимо внести корректировку'
                    a_comm = a_comm + prom_comm + ''
                    comp_table[i] = True
            else:
                prom_comm = str(serv_32[i][0]) + ' не заказывается Потребителем услуг. Необходимо внести корректировку'
                a_comm = a_comm + prom_comm + ''
                comp_table[i] = True

        for i in range(0, len(serv_32)):
            if comp_table[i] != True:
                if ('3.2.1' in str(serv_32[i][0]) or '3.2.2' in str(serv_32[i][0]) or '3.2.3' in str(
                        serv_32[i][0])) and ip_addr == []:
                    prom_comm = 'Для заказа ' + str(
                        serv_32[i][0]) + ' требуется заказ публичных ip-адресов. Без них заказ данной услуги невозможен'
                    a_comm = a_comm + prom_comm + ''
                    comp_table[i] = True
                else:
                    prom_comm = 'Для ' + str(serv_32[i][0]) + ' не найдены критерии проверки. Прошу перепроверить.'
    else:
        if vms != None and vms != 0:
            comp_vms = None
            sec_vms_1 = vms // 50
            if vms % 50 != 0:
                sec_vms_1 += 1
            prom_comm = 'Не заказана Услуга по мониторингу, выявлению и активному реагированию на инциденты информационной безопасности на конечных точках (услуга 3.2.5)' + '. Для заказываемого количества ВМ - ' + str(
                vms) + ' необходимо заказывать - ' + str(
                sec_vms_1) + ' штук данной услуги. Каждая единица услуги имеет квант 50 ВМ.'
            a_comm = a_comm + prom_comm + ''
            comp_vms = True
        if ip_addr != []:
            prom_comm = 'Не заказана Услуга анализа внешнего периметра на наличие уязвимостей. (услуга 3.2.3)' + '. Для заказываемого количества  - ' + str(
                ip_addr[0][0]) + ' необходимо заказывать - ' + str(
                ip_addr[0][1]) + ' штук данной услуги по количеству заказываемых ip-адресов.'
            a_comm = a_comm + prom_comm + ''
    if len(a_comm) == start_len or a_comm.strip().endswith("Таблица 20"):
        a_comm += "Замечаний по проверке ИБ не выявлено.\n"
    _log('Завершена проверка поступившего файла на услуги ИБ')
    return a_comm



# NOTE:
# sec_sec() в оригинале продолжается дальше (проверки сопровождения, 3.2.*, ip и т.д.).
# Для корректной работы эти части также должны быть в файле.
# Мы подставим их из ранее собранного security_merged.py, чтобы 1:1 сохранить поведение.




# -----------------------------
# Web entry point (для Django)
# -----------------------------

def reset_state():
    """Сбрасывает глобальные накопленные данные перед новым запуском."""
    global a_comm, prom_comm
    a_comm = 'Результаты обработки поступивших данных\n'
    prom_comm = ''
    # удалить ранее распарсенные таблицы заказа
    for k in list(globals().keys()):
        if k.startswith('new_data_serv_'):
            globals().pop(k, None)


def run_web(order_excel_path: str) -> str:
    """Запуск проверки ИБ для сайта.

    Принимает путь к Excel-файлу заказа и возвращает текст отчёта.
    Никаких input()/print() и интерактива.
    """
    reset_state()

    # 1) читаем заказ
    qs = excel_read('new', order_excel_path)
    if qs == 7:
        return 'Данный файл не содержит листов, указанных в конфигурационном файле'

    # 2) проверка
    return sec_sec()
