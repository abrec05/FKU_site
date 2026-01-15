#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sravn_web.py

WEB-версия проверки "Сравнение заказов" (из sravn.py):
- Никаких input()/print()
- Никакой записи файлов на диск (возвращаем текст отчёта)
- Путь к конфигам file_config.txt и file_config_art.txt берётся относительно этого файла

Использование в Django:
    from ost_zak import sravn_web
    report = sravn_web.run_web(new_excel_path, old_excel_path)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, List, Tuple

import numpy as np
from openpyxl import load_workbook

# ------------------------ утилиты (из исходника) ------------------------

def r_file_conf(filename: str) -> List[List[List[str]]]:
    """
    Считывает текстовый файл построчно и возвращает массив,
    где каждый элемент - это [левая_часть_до_&, правая_часть_после_&], разделение по ';'.
    """
    with open(filename, "r", encoding="utf-8") as file:
        lines = file.readlines()

    new_lines = [(line.strip()).split(';') for line in lines]
    res: List[List[List[str]]] = []
    for row in new_lines:
        if '&' not in row:
            continue
        res.append([row[:row.index('&')], row[row.index('&') + 1:]])
    return res

def _col_name(names_col, j: int) -> str:
    try:
        return str(names_col[j]).strip()
    except Exception:
        return f"Колонка {j}"
def check_range(r: List[str], i: int) -> bool:
    """Проверяет, попадает ли i в список значений/диапазонов вида ["1", "2-5", ...]."""
    for t in r:
        if '-' in t:
            start, end = map(int, t.split("-"))
            if start <= i <= end:
                return True
        else:
            if i == int(t):
                return True
    return False


# ------------------------ конфигурация чтения Excel (вырезка) ------------------------

def load_excel_read_config(config_path: str = "file_config.txt") -> dict:
    """
    Парсит file_config.txt так же, как в исходнике. Возвращает словарь с настройками,
    которые нужны для excel_read().
    """
    lines = 0
    sh_line = 0
    sh_n_line = 0

    t_n_s: List[str] = []
    u_t_n: List[str] = []
    u_p_t: List[str] = []
    t_t_n: List[str] = []
    t_t_ch: List[str] = []
    t_t_p: List[str] = []

    table_number_line = 1000000
    unic_table_number_line = 1000000
    unic_params_table_line = 1000000
    tip_table_number_line = 1000000
    tip_table_ch_line = 1000000
    tip_params_table_line = 1000000

    table_number: List[int] = []
    unic_table_number: List[int] = []
    unic_params_table: List[list] = []
    tip_table_number: List[int] = []
    tip_table_ch: List[Any] = []
    tip_params_table: List[Any] = []

    prom_unic_pt = 0
    tt = 10**13
    named_table = None

    with open(config_path, 'r', encoding='utf-8') as f:
        for _ in f:
            lines += 1

        f.seek(0)
        for i in range(lines):
            prom_str = f.readline()
            if 'sheet_number' in prom_str:
                sh_line = i + 1
            elif 'sheet_names' in prom_str:
                sh_n_line = i + 1

        f.seek(0)
        sheet_number = None
        sheet_names = None
        for i in range(lines):
            prom_str = f.readline()
            if i == sh_line:
                sheet_number = int(prom_str)
            elif i == sh_n_line:
                sheet_names = prom_str.split(';	')
                sheet_names = np.array(sheet_names)
                for j in range(len(sheet_names)):
                    if '\n' in str(sheet_names[j]):
                        s = str(sheet_names[j])
                        sheet_names[j] = s[:s.find('\n')]

        if sheet_number is None or sheet_names is None:
            raise RuntimeError("Не удалось прочитать sheet_number / sheet_names из file_config.txt")

        f.seek(0)
        for j in range(int(sheet_number)):
            f.seek(0)
            for _ in range(lines):
                nn = f'sheet_{j+1}'
                prom_str = f.readline()
                if nn in prom_str:
                    t_n_s.append(f'1table_number_{j+1}')
                    u_t_n.append(f'unic_table_number_{j+1}')
                    u_p_t.append(f'unic_params_table_{j+1}')
                    t_t_n.append(f'tip_table_number_{j+1}')
                    t_t_ch.append(f'tip_table_ch_{j+1}')
                    t_t_p.append(f'tip_params_table_{j+1}')

        f.seek(0)
        for j in range(int(sheet_number)):
            table_number_line = unic_table_number_line = unic_params_table_line = 1000000
            tip_table_number_line = tip_table_ch_line = tip_params_table_line = 1000000
            f.seek(0)

            for i in range(lines):
                prom_str = f.readline()
                if t_n_s[j] in prom_str:
                    table_number_line = i + 1
                elif u_t_n[j] in prom_str:
                    unic_table_number_line = i + 1
                elif u_p_t[j] in prom_str:
                    unic_params_table_line = i + 1
                elif t_t_n[j] in prom_str:
                    tip_table_number_line = i + 1
                elif t_t_ch[j] in prom_str:
                    tip_table_ch_line = i + 1
                elif t_t_p[j] in prom_str:
                    tip_params_table_line = i + 1

                if i == table_number_line:
                    table_number.append(int(prom_str))
                elif i == unic_table_number_line:
                    unic_table_number.append(int(prom_str))
                    prom_unic_pt = int(prom_str)
                elif i >= unic_params_table_line and prom_unic_pt != 0:
                    unic_params_table.append(prom_str.split(';	'))
                    prom_unic_pt -= 1
                elif i == tip_table_number_line:
                    tip_table_number.append(int(prom_str))
                elif i == tip_table_ch_line:
                    tip_table_ch.append(prom_str.split(';	'))
                elif i == tip_params_table_line:
                    tip_params_table.append(prom_str.split(';	'))

        f.seek(0)
        for i in range(lines):
            prom_str = f.readline()
            if 'named_table' in prom_str:
                tt = i + 1
            if i == tt:
                named_table = prom_str

    if named_table is None:
        raise RuntimeError("Не удалось прочитать named_table из file_config.txt")

    if '\n' in named_table:
        named_table = named_table[:named_table.find('\n')]

    for i in range(len(unic_params_table)):
        for j in range(len(unic_params_table[i])):
            s = str(unic_params_table[i][j])
            if '\n' in s:
                unic_params_table[i][j] = s[:s.find('\n')]

    for i in range(len(tip_params_table)):
        for j in range(len(tip_params_table[i])):
            s = str(tip_params_table[i][j])
            if '\n' in s:
                tip_params_table[i][j] = s[:s.find('\n')]

    tt_list: List[Any] = []
    for i in range(len(tip_table_ch)):
        if tip_table_ch[i] != []:
            for j in range(len(tip_table_ch[i])):
                if '-' in str(tip_table_ch[i][j]):
                    dd = str(tip_table_ch[i][j]).split('-')
                    for h in range(int(dd[0]), int(dd[1])):
                        tt_list.append(h)
                elif '' not in str(tip_table_ch[i][j]) and None not in str(tip_table_ch[i][j]):
                    tt_list.append(tip_table_ch[i][j])
    tip_table_ch = tt_list

    return {
        "sheet_number": int(sheet_number),
        "sheet_names": sheet_names,
        "table_number": table_number,
        "unic_params_table": unic_params_table,
        "tip_table_ch": tip_table_ch,
        "tip_params_table": tip_params_table,
        "named_table": named_table,
    }


# ------------------------ чтение Excel в глобальные таблицы (как в исходнике) ------------------------

CFG: dict = {}
sheet_number: int
sheet_names: Any
table_number: List[int]
unic_params_table: List[list]
tip_table_ch: List[Any]
tip_params_table: List[Any]
named_table: str


def _init_config(config_dir: str | Path) -> None:
    global CFG, sheet_number, sheet_names, table_number, unic_params_table, tip_table_ch, tip_params_table, named_table
    config_dir = Path(config_dir)
    CFG = load_excel_read_config(str(config_dir / "file_config.txt"))
    sheet_number = CFG["sheet_number"]
    sheet_names = CFG["sheet_names"]
    table_number = CFG["table_number"]
    unic_params_table = CFG["unic_params_table"]
    tip_table_ch = CFG["tip_table_ch"]
    tip_params_table = CFG["tip_params_table"]
    named_table = CFG["named_table"]


def _clear_loaded_tables() -> None:
    to_del = [k for k in list(globals().keys()) if k.startswith("old_data_serv_") or k.startswith("new_data_serv_")]
    for k in to_del:
        del globals()[k]


def excel_read(new_old: str, zak: str) -> int:
    """
    Читает Excel и складывает данные в globals() в переменные:
        new_data_serv_<лист>_<таблица>   или old_data_serv_<лист>_<таблица>
    """
    zk_type = new_old
    qs = 0

    wb = load_workbook(filename=zak, data_only=True)
    sheets = wb.sheetnames
    ws = []

    for sh in sheets:
        for j in range(0, len(sheet_names)):
            if str(sheet_names[j]) in sh:
                ws.append(wb[sh])

    if ws == []:
        qs = 7

    for i in range(0, len(ws)):
        rows = ws[i].max_row
        columns = ws[i].max_column

        tables_rows = []
        for ii in range(1, rows):
            for jj in range(1, columns):
                if str(named_table) in str(ws[i].cell(row=ii, column=jj).value):
                    tables_rows.append(ii)

        for k in range(0, len(tables_rows)):
            col_params: List[int] = []
            row_params = None
            chek_l = 10**13
            name = None

            end_row = tables_rows[k + 1] if k != len(tables_rows) - 1 else rows

            name = None

            for ii in range(tables_rows[k], end_row):
                new_line: List[Any] = []
                for jj in range(1, columns):
                    if named_table in str(ws[i].cell(row=ii, column=jj).value):
                        name = str(ws[i].cell(row=ii, column=jj).value)
                        nn = name.find(named_table)
                        if name[nn + len(named_table) + 2] != '.':
                            name = name[nn + len(named_table) + 1: nn + len(named_table) + 3]
                        else:
                            name = name[nn + len(named_table) + 1]

                    for m in range(0, len(unic_params_table)):
                        if name is not None and int(name) == int(unic_params_table[m][0]):
                            chek_l = m
                    if chek_l >= len(unic_params_table):
                        chek_l = -1

                    if chek_l != -1 and chek_l != len(unic_params_table):
                        for m in range(1, len(unic_params_table[chek_l])):
                            if str(unic_params_table[chek_l][m]) in str(ws[i].cell(row=ii, column=jj).value):
                                if jj not in col_params:
                                    col_params.append(jj)
                                    row_params = ii
                        for m in range(0, len(col_params)):
                            if jj == int(col_params[m]) and row_params is not None and ii > row_params:
                                new_line.append(ws[i].cell(row=ii, column=jj).value)
                    else:
                        for m in tip_table_ch:
                            if name is not None and int(name) == int(m):
                                chek_l = m
                        if chek_l != 10**13:
                            for m in range(0, len(tip_params_table[i])):
                                if str(tip_params_table[i][m]) in str(ws[i].cell(row=ii, column=jj).value):
                                    if jj not in col_params:
                                        col_params.append(jj)
                                        row_params = ii
                            for m in range(0, len(col_params)):
                                if jj == int(col_params[m]) and row_params is not None and ii > row_params:
                                    new_line.append(ws[i].cell(row=ii, column=jj).value)

                if name is None:
                    continue

                q = f"{i+1}_{name}"
                var_name = f"{zk_type}_data_serv_{q}"

                if var_name not in globals():
                    if new_line != [] and len(new_line) > 1 and new_line[1] is not None and new_line[0] is not None:
                        globals()[var_name] = []
                        globals()[var_name].append(new_line)
                else:
                    if new_line != [] and len(new_line) > 1 and new_line[1] is not None and new_line[0] is not None:
                        globals()[var_name].append(new_line)

    wb.close()
    return qs


# ------------------------ СРАВНЕНИЕ (проверка №2) ------------------------

def sravn(listik: int, tabl: int, pr_com: List[str], names_col: List[Any]) -> List[str]:
    """
    Сравнивает old_data_serv_<лист>_<таблица> vs new_data_serv_<лист>_<таблица>
    и дописывает человеко-читаемые комментарии в pr_com.
    """
    tables = [
        'Услуги 1-2.1',
        'Услуги 2.2-2.19',
        'Услуги 3-4'
    ]

    fl = False
    pr_com.append(None)
    pr = pr_com
    str_F = []

    old_key = f'old_data_serv_{listik}_{tabl}'
    new_key = f'new_data_serv_{listik}_{tabl}'
    if old_key in globals() and new_key in globals():
        old = globals()[old_key]
        new = globals()[new_key]
    else:
        pr_com[pr_com.index(None)] = f'Таблица {listik}_{tabl} отсутствует в одном из файлов (new/old).'
        return pr_com

    for i in names_col:
        if check_range(i[0], tabl):
            names_col = i[1]
            break

    if new != old:
        n = 0
        row = 1
        while row < len(old):

            if fl:
                pr_com.append('Строка ' + str(old[row][0]) + ' отсутствует в новом файле c параметрами:\n')
                for _ in range(row, len(old)):
                    str_F.append(old[row])
                    for values in range(0, len(old[row]) - 1):
                        pr_com.append(f'{names_col[values]}: {old[row][values]}\n')
                break

            if not fl:
                if not (type(old[row][0]) is int):
                    if str(old[row][0]).count('.') == 2:
                        nomer = str(old[row][0]).split('.')
                        nomer = list(map(float, nomer))
                        nomer_old = nomer[0] + nomer[1] / 10 + nomer[2] / 100
                    else:
                        nomer_old = float(old[row][0])
                else:
                    nomer_old = float(old[row][0])

                if not (type(new[row + n][0]) is int):
                    if str(new[row + n][0]).count('.') == 2:
                        nomer = str(new[row + n][0]).split('.')
                        nomer = list(map(float, nomer))
                        nomer_new = nomer[0] + nomer[1] / 10 + nomer[2] / 100
                    else:
                        nomer_new = float(new[row + n][0])
                else:
                    nomer_new = float(new[row + n][0])

                if old[row][0] != new[row + n][0] and nomer_old > nomer_new and not fl:
                    n += 1
                    if len(new) - 1 < row + n:
                        fl = True
                    continue
                elif old[row][0] != new[row + n][0]:
                    str_F.append(old[row])
                    pr_com.append(
                        'Строка ' + str(old[row][0]) + ' старого файла отсутствует в новом файле c параметрами:\n'
                    )
                    for values in range(0, min(23, len(old[row]), len(names_col))):
                        pr_com.append(f'{names_col[values]}: {old[row][values]}\n')
                    n -= 1
                    row += 1
                    continue

            if old[row] != new[row + n]:
                f = True
                t = 0
                if len(old[row]) != len(new[row + n]):
                    t = 1
                for j in range(0, len(old[row]) - t):
                    if str(old[row][j]) != str(new[row + n][j]) and j != 5:
                        if len(old[row]) > 30:
                            if j != 32:
                                if f:
                                    pr_com.append('Строка ' + str(old[row][0]) + ' :' + '\n')
                                    f = False
                                pr_com.append(
                                    'Вероятно допущена ошибка в параметре ' + _col_name(names_col, j)
                                    + '. В старом заказе данный параметр был ' + str(old[row][j])
                                    + ', а в новом ' + str(new[row + n][j]) + '.' + '\n'
                                )
                                f = True
                        elif j != 12 and old[row][j] != new[row + n][j]:
                            if f:
                                pr_com.append('Строка ' + str(old[row][0]) + ' :' + '\n')
                                f = False
                            pr_com.append(
                                'Вероятно допущена ошибка в параметре ' + _col_name(names_col, j)
                                + '. В старом заказе данный параметр был ' + str(old[row][j])
                                + ', а в новом ' + str(new[row + n][j]) + '.' + '\n'
                            )

            row += 1

        if pr != pr_com:
            pr_com[pr_com.index(None)] = (
                'Количество НЕ новых строк на листе ' + tables[listik - 1]
                + ' в таблице № ' + str(tabl)
                + ' не совпало с количеством ранее указанных в старом заказе. Количество строк ='
                + str(len(old) + n) + '\n'
            )
        else:
            pr_com[pr_com.index(None)] = (
                'Количество НЕ новых строк на листе ' + tables[listik - 1]
                + ' в таблице № ' + str(tabl)
                + ' совпало с количеством ранее указанных в старом заказе. Количество строк ='
                + str(len(old)) + '\n'
            )
    else:
        pr_com[pr_com.index(None)] = (
            'Количество НЕ новых строк на листе ' + tables[listik - 1]
            + ' в таблице № ' + str(tabl)
            + ' совпало с количеством ранее указанных в старом заказе. Количество строк ='
            + str(len(old)) + '\n'
        )
    return pr_com


def run_web(new_excel_path: str, old_excel_path: str, config_dir: str | None = None) -> str:
    """
    WEB-точка входа.
    Возвращает текст отчёта сравнения (без записи файлов на диск).

    config_dir: папка, где лежат file_config.txt и file_config_art.txt.
                Если None — берём папку этого файла (sravn_web.py).
    """
    base = Path(config_dir) if config_dir else Path(__file__).resolve().parent

    _init_config(base)
    names_cfg = r_file_conf(str(base / "file_config_art.txt"))

    start_values: List[Tuple[int, int]] = [
        (1, 1), (1, 2),
        (2, 3), (2, 4), (2, 5), (2, 6), (2, 7), (2, 8), (2, 9), (2, 10), (2, 11), (2, 12), (2, 13), (2, 14),
        (2, 15), (2, 16), (2, 17), (2, 18), (2, 19),
        (3, 20), (3, 21), (3, 22), (3, 23),
    ]

    _clear_loaded_tables()

    qs_new = excel_read('new', str(Path(new_excel_path)))
    qs_old = excel_read('old', str(Path(old_excel_path)))

    if qs_new == 7:
        return "Новый заказ: файл не содержит листов, указанных в конфигурационном файле"
    if qs_old == 7:
        return "Старый заказ: файл не содержит листов, указанных в конфигурационном файле"

    comments: List[str] = []
    for a, b in start_values:
        comments = sravn(a, b, comments, names_cfg)

    return "\n".join(str(x) for x in comments if x is not None)
